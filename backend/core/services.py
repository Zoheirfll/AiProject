import datetime
import os
import re
import shutil

import openpyxl
from django.core.files import File
from django.core.mail import EmailMessage, EmailMultiAlternatives

from employees.models import Employee


def _strip_html(html):
    return re.sub(r"<[^>]+>", " ", html).strip()


def send_mail_log(mail_log, destinataires, cc=None, bcc=None):
    """Send a MailLog's subject/body, respecting its format (US-E4-02):
    TEXTE sends a plain-text message; HTML sends a real HTML email with a
    plain-text fallback (for clients that don't render HTML).
    """
    if mail_log.format == "HTML":
        message = EmailMultiAlternatives(
            subject=mail_log.subject,
            body=_strip_html(mail_log.body),
            to=destinataires,
            cc=cc or None,
            bcc=bcc or None,
        )
        message.attach_alternative(mail_log.body, "text/html")
        message.send(fail_silently=False)
    else:
        EmailMessage(
            subject=mail_log.subject,
            body=mail_log.body,
            to=destinataires,
            cc=cc or None,
            bcc=bcc or None,
        ).send(fail_silently=False)

REQUIRED_COLUMNS = ["matricule", "nom", "prenom"]
COLUMN_MAP = {
    "matricule": "matricule",
    "nom": "nom",
    "prenom": "prenom",
    "email": "email",
    "departement": "departement",
    "département": "departement",
    "poste": "poste",
    "categorie": "categorie",
    "catégorie": "categorie",
    "num_contrat": "num_contrat",
    "n° contrat": "num_contrat",
    "date_embauche": "date_embauche",
    "date d'embauche": "date_embauche",
    "date_fin_contrat": "date_fin_contrat",
    "date fin contrat": "date_fin_contrat",
}
DATE_FIELDS = {"date_embauche", "date_fin_contrat"}
TEXT_FIELDS = ["nom", "prenom", "email", "departement", "poste", "categorie", "num_contrat"]

DATE_FORMATS = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]


def _coerce_date(value):
    """Return a date, or raise ValueError with a French message."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"date invalide: '{text}'")


def build_field_map(header, mapping_override=None):
    """Map spreadsheet column index -> system field name.

    mapping_override, if given, is {system_field: excel_column_header}
    (as saved via ImportConfig) and takes priority over the built-in
    header aliases in COLUMN_MAP.
    """
    field_by_col = {}
    if mapping_override:
        header_to_field = {
            str(col).strip().lower(): field
            for field, col in mapping_override.items()
            if col
        }
        for i, h in enumerate(header):
            if h in header_to_field:
                field_by_col[i] = header_to_field[h]
    else:
        for i, h in enumerate(header):
            if h in COLUMN_MAP:
                field_by_col[i] = COLUMN_MAP[h]
    return field_by_col


def parse_employee_excel(file_obj, mapping_override=None):
    """Parse an uploaded Excel file into Employee rows.

    mapping_override: optional {system_field: excel_column_header} dict
    (from ImportConfig.mapping) to use instead of the built-in COLUMN_MAP.

    Columns that don't match a known system field are NOT ignored — HR data
    is wide open (leave, training, evaluations, disciplinary notes, anything)
    and the app shouldn't hardcode an exhaustive list of what an Excel sheet
    is allowed to contain. Any such column is kept verbatim per employee in
    Employee.donnees_supplementaires, and also fed to the analyste agent
    (agents.analyste) so it can reason about whatever the sheet is actually
    about instead of assuming it's always contracts.

    Returns (total, imported, errors, lignes) where errors is a list of
    {"ligne": int, "message": str} dicts and lignes is a list of per-row
    dicts (system fields + extra columns) for successfully imported rows.
    """
    workbook = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0, 0, [{"ligne": 0, "message": "Fichier vide"}], []

    header = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    field_by_col = build_field_map(header, mapping_override)
    extra_cols = {i: h for i, h in enumerate(header) if i not in field_by_col and h}

    missing = [c for c in REQUIRED_COLUMNS if c not in field_by_col.values()]
    if missing:
        return 0, 0, [
            {"ligne": 1, "message": f"Colonnes manquantes: {', '.join(missing)}"}
        ], []

    total = 0
    imported = 0
    errors = []
    lignes = []

    for line_num, row in enumerate(rows[1:], start=2):
        if all(cell is None for cell in row):
            continue
        total += 1
        data = {field_by_col[i]: row[i] for i in field_by_col if i < len(row)}
        extra = {
            extra_cols[i]: row[i]
            for i in extra_cols
            if i < len(row) and row[i] not in (None, "")
        }

        if not data.get("matricule") or not data.get("nom") or not data.get("prenom"):
            errors.append({"ligne": line_num, "message": "Champs requis manquants"})
            continue

        try:
            defaults = {
                field: str(data.get(field) or "").strip() for field in TEXT_FIELDS
            }
            for field in DATE_FIELDS:
                defaults[field] = _coerce_date(data.get(field))
            defaults["donnees_supplementaires"] = {k: str(v) for k, v in extra.items()}

            Employee.objects.update_or_create(
                matricule=str(data["matricule"]).strip(),
                defaults=defaults,
            )
            imported += 1
            lignes.append({**defaults, "matricule": str(data["matricule"]).strip()})
        except Exception as exc:  # noqa: BLE001
            errors.append({"ligne": line_num, "message": str(exc)})

    return total, imported, errors, lignes


MAIL_COLUMN_MAP = {
    "nom": "nom",
    "email": "email",
    "sujet": "sujet",
}


def parse_mail_masse_excel(file_obj):
    """Parse an uploaded Excel file into ad-hoc mail recipients.

    Expected columns: email (required), nom (optional), sujet (optional —
    falls back to a default subject supplied by the caller when blank).

    Returns (rows, errors) where rows is a list of {"nom", "email", "sujet"}
    dicts and errors is a list of {"ligne": int, "message": str} dicts.
    """
    workbook = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = workbook.active

    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return [], [{"ligne": 0, "message": "Fichier vide"}]

    header = [str(cell).strip().lower() if cell else "" for cell in all_rows[0]]
    field_by_col = {i: MAIL_COLUMN_MAP[h] for i, h in enumerate(header) if h in MAIL_COLUMN_MAP}

    if "email" not in field_by_col.values():
        return [], [{"ligne": 1, "message": "Colonne manquante: email"}]

    rows = []
    errors = []

    for line_num, row in enumerate(all_rows[1:], start=2):
        if all(cell is None for cell in row):
            continue
        data = {field_by_col[i]: row[i] for i in field_by_col if i < len(row)}

        email = str(data.get("email") or "").strip()
        if not email:
            errors.append({"ligne": line_num, "message": "Email manquant"})
            continue

        rows.append(
            {
                "nom": str(data.get("nom") or "").strip(),
                "email": email,
                "sujet": str(data.get("sujet") or "").strip(),
            }
        )

    return rows, errors


def scan_dossier_surveille():
    """Scheduler tick: look for new Excel files in ImportConfig.dossier_surveille,
    import each one, move it to a /processed subfolder, and notify over WebSocket.
    """
    from .models import ExcelImport, ImportConfig

    config = ImportConfig.get_solo()
    dossier = config.dossier_surveille
    if not dossier or not os.path.isdir(dossier):
        return

    processed_dir = os.path.join(dossier, "processed")
    os.makedirs(processed_dir, exist_ok=True)

    for name in sorted(os.listdir(dossier)):
        if not name.lower().endswith((".xlsx", ".xls")):
            continue
        path = os.path.join(dossier, name)
        if not os.path.isfile(path):
            continue

        with open(path, "rb") as fh:
            excel_import = ExcelImport.objects.create(
                fichier=File(fh, name=name),
                nom_fichier_origine=name,
                source=ExcelImport.Source.DOSSIER,
            )

        lignes = []
        try:
            with excel_import.fichier.open("rb") as fh:
                total, imported, errors, lignes = parse_employee_excel(fh, config.mapping or None)
            excel_import.lignes_total = total
            excel_import.lignes_importees = imported
            excel_import.lignes_erreurs = len(errors)
            excel_import.erreurs = errors
            excel_import.status = (
                ExcelImport.Status.SUCCESS
                if imported > 0 or (total == 0 and not errors)
                else ExcelImport.Status.FAILED
            )
        except Exception as exc:  # noqa: BLE001
            excel_import.status = ExcelImport.Status.FAILED
            excel_import.erreurs = [{"ligne": 0, "message": str(exc)}]
        excel_import.save()

        if excel_import.status == ExcelImport.Status.SUCCESS:
            from agents.analyste import analyser_import

            try:
                analyser_import(excel_import, lignes=lignes)
            except Exception:  # noqa: BLE001
                pass  # best-effort — never let the analyste agent break folder-watch imports

        shutil.move(path, os.path.join(processed_dir, name))

        from integrations.notifications import notify

        notify(
            {
                "type": "import",
                "id": excel_import.id,
                "fichier": name,
                "status": excel_import.status,
                "lignes_importees": excel_import.lignes_importees,
            }
        )
